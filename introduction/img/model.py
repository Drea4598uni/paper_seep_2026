"""
Computational-cost model for 24 wind-turbine / wind-farm CFD cases.
All assumptions documented in the accompanying report.
Outputs: wind_cfd_cost_grid.xlsx and cases.json (for the report table).
"""
import json, math
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ----------------------------------------------------------------------
# 1. ASSUMPTIONS / INPUTS
# ----------------------------------------------------------------------

# --- Mesh model (cells in MILLIONS) ---
# Per-turbine cell count (5 MW baseline)
PER_TURBINE = {
    ("FG", "URANS"): 60.0,    # blade-resolved, RANS-type wall treatment
    ("FG", "LES"):   350.0,   # blade-resolved, ABL-LES, near-wall + wake
    ("AL", "URANS"): 6.0,     # actuator line/disk refinement zone
    ("AL", "LES"):   15.0,    # actuator line/disk refinement zone
}
# Background ABL / domain mesh (millions), depends on domain size (# turbines)
BACKGROUND = {
    ("single", "URANS"): 4.0,   ("single", "LES"): 40.0,
    ("farm20", "URANS"): 20.0,  ("farm20", "LES"): 200.0,
    ("farm40", "URANS"): 35.0,  ("farm40", "LES"): 350.0,
}
# Turbine-size scaling applied to per-turbine cells (higher Re, larger domain)
SIZE_FACTOR = {
    ("FG", 5): 1.0,  ("FG", 15): 1.8,
    ("AL", 5): 1.0,  ("AL", 15): 1.4,
}

# --- Rotor diameter (m) and minimum cell size in the MAX-refinement zone ---
ROTOR_D = {5: 126.0, 15: 240.0}    # NREL 5 MW / IEA 15 MW
# Full geometry: smallest cell = wall-normal first layer (y+ ~ 1), in metres
FG_DX_MIN = {"LES": 1.0e-5, "URANS": 2.0e-5}
# Actuator line/disk: smallest cell = rotor/wake refinement box, as fraction of D
AL_DX_OVER_D = {"LES": 1.0/100.0, "URANS": 1.0/64.0}

# --- Wind-farm layout (aligned rectangular array, wind along columns) ---
# columns (streamwise) x rows (lateral); spacing Sx (streamwise) x Sy (lateral) in D
LAYOUT = {
    "single": ("1×1", "-"),
    "farm20": ("5×4", "7×5"),
    "farm40": ("8×5", "7×5"),
}

# --- Computational domain extent (Lx x Ly x Lz) in rotor diameters ---
# (streamwise x lateral x vertical); same outer domain for FG and AL.
# Consistent with the aligned arrays above: array span + ~6 D upstream and
# ~10-15 D downstream (streamwise) and ~5 D side margins (lateral).
DOMAIN_D = {
    "single": (24, 8, 6),
    "farm20": (44, 26, 6),
    "farm40": (64, 32, 6),
}

# --- Time / convergence model -------------------------------------------
# one wind speed, neutral ABL. Physical simulated time T_sim is set by the
# dominant convergence criterion; #steps = T_sim / dt.
U_HUB = 10.0                       # hub-height inflow velocity (m/s), one wind speed
RPM   = {5: 12.0, 15: 7.5}         # NREL 5 MW / IEA 15 MW near rated

# Blade-resolved (full geometry) -> converge rotor LOADS over revolutions
N_REV_FG = {"URANS": 30, "LES": 15}
AZIM_FG_URANS_DEG = 2.0            # URANS time step = 2 deg azimuth
DT_FG_LES = 1.0e-4                 # s, CFL-limited near-blade LES time step

# Actuator line/disk -> converge WAKE/POWER statistics over flow-through times
N_FT_AL = {
    ("LES", "single"): 10, ("LES", "farm20"): 12, ("LES", "farm40"): 14,
    ("URANS", "single"): 6, ("URANS", "farm20"): 8, ("URANS", "farm40"): 10,
}
# Actuator time step: actuator tip advances <= 1 finest cell per step
# dt_AL = dx_min / V_tip  (computed per case)

# --- Solver cost: core-seconds per cell per time step ---
# C_URANS anchored to blade-resolved URANS (~16.3k CPUh, 55 M cells, 5400 steps);
# C_LES lower (explicit/PISO LES) and calibrated so converged actuator-LES farm
# runs fall in the published "days-to-weeks on O(1e3-1e4) cores" range.
C_SOLVER = {"URANS": 2.0e-4, "LES": 3.0e-5}

# --- Leonardo (CINECA) hardware ---
DCGP_CORES_PER_NODE = 112      # 2 x Intel Xeon 8480+ (56c)
DCGP_NODE_KW        = 1.1
BOOSTER_NODE_KW     = 2.5      # 4x A100 64GB + 32c Ice Lake
GPU_NODE_CORE_EQUIV = 600      # 1 Booster node ~ 600 CPU cores for FV-CFD

# Job size (nodes) used for wall-clock estimate
DCGP_NODES = {"single": 16, "farm": 64}
BOOSTER_NODES = {("AL"): 32, ("FG_single"): 64, ("FG_farm"): 256}

# --- Energy / cost ---
PUE = 1.2
EUR_PER_KWH = 0.15

# ----------------------------------------------------------------------
# 2. CASE DEFINITIONS
# ----------------------------------------------------------------------
CONFIGS = [
    ("1", "Single turbine - full geometry",        "FG", "single", 1),
    ("2", "Single turbine - actuator line/disk",   "AL", "single", 1),
    ("3", "Wind farm 20 - full geometry",          "FG", "farm20", 20),
    ("4", "Wind farm 40 - full geometry",          "FG", "farm40", 40),
    ("5", "Wind farm 20 - actuator line/disk",     "AL", "farm20", 20),
    ("6", "Wind farm 40 - actuator line/disk",     "AL", "farm40", 40),
]
MODELS = ["LES", "URANS"]
SIZES  = [5, 15]

def domain_key(domain):
    return "single" if domain == "single" else "farm"

rows = []
for cfg_id, cfg_name, mtype, domain, nturb in CONFIGS:
    for model in MODELS:
        for size in SIZES:
            dk = domain_key(domain)
            # cells (millions)
            cells_M = (nturb * PER_TURBINE[(mtype, model)] * SIZE_FACTOR[(mtype, size)]
                       + BACKGROUND[(domain, model)])
            cells = cells_M * 1e6

            # --- geometry / refinement ---
            D = ROTOR_D[size]
            if mtype == "FG":
                dx_min = FG_DX_MIN[model]
                dx_over_D = dx_min / D
            else:
                dx_over_D = AL_DX_OVER_D[model]
                dx_min = dx_over_D * D
            dxd, dyd, dzd = DOMAIN_D[domain]
            Lx, Ly, Lz = dxd * D, dyd * D, dzd * D
            domain_km = f"{Lx/1000:.1f}×{Ly/1000:.1f}×{Lz/1000:.1f}"
            domain_D = f"{dxd}×{dyd}×{dzd}"
            layout, spacing = LAYOUT[domain]

            # --- physical time, time step, #steps, flow-throughs, revolutions ---
            rpm = RPM[size]
            T_rev = 60.0 / rpm                       # s per rotor revolution
            V_tip = math.pi * D * rpm / 60.0         # blade-tip speed (m/s)
            T_ft  = Lx / U_HUB                        # one flow-through time (s)
            # ALL cases must develop AND statistically average the WAKE, so every
            # case (blade-resolved included) is run for several flow-through times.
            nft_target = N_FT_AL[(model, domain)]
            T_sim = nft_target * T_ft
            if mtype == "FG":
                # blade-resolved time step (small): 2 deg azimuth (URANS) / CFL (LES)
                dt = (AZIM_FG_URANS_DEG/360.0)*T_rev if model == "URANS" else DT_FG_LES
            else:
                dt = dx_min / V_tip                   # actuator-tip transit limit
            steps = int(math.ceil(T_sim / dt))
            flow_throughs = T_sim / T_ft
            rotor_revs = T_sim / T_rev

            # core-hours (CPU-core-hours)
            core_h = cells * steps * C_SOLVER[model] / 3600.0
            # map to Leonardo allocation
            if model == "URANS":
                part = "DCGP (CPU)"
                node_h = core_h / DCGP_CORES_PER_NODE
                node_kw = DCGP_NODE_KW
                nodes = DCGP_NODES[dk]
            else:  # LES on Booster GPU
                part = "Booster (GPU)"
                node_h = core_h / GPU_NODE_CORE_EQUIV
                node_kw = BOOSTER_NODE_KW
                if mtype == "FG":
                    nodes = BOOSTER_NODES["FG_single"] if dk == "single" else BOOSTER_NODES["FG_farm"]
                else:
                    nodes = BOOSTER_NODES["AL"]
            wall_h = node_h / nodes
            energy_kwh = node_h * node_kw * PUE
            cost_eur = energy_kwh * EUR_PER_KWH
            rows.append({
                "case": f"{cfg_id}",
                "config": cfg_name,
                "model": model,
                "size_MW": size,
                "rotor_D_m": D,
                "dx_min_m": dx_min,
                "dx_over_D": dx_over_D,
                "domain_km": domain_km,
                "domain_D": domain_D,
                "layout": layout,
                "spacing_D": spacing,
                "T_sim_s": T_sim,
                "flow_throughs": flow_throughs,
                "rotor_revs": rotor_revs,
                "n_turbines": nturb,
                "cells_M": round(cells_M, 1),
                "n_steps": steps,
                "core_hours": core_h,
                "partition": part,
                "node_hours": node_h,
                "nodes": nodes,
                "wall_days": wall_h / 24.0,
                "energy_MWh": energy_kwh / 1000.0,
                "cost_EUR": cost_eur,
            })

# ----------------------------------------------------------------------
# 3. CONSOLE SUMMARY
# ----------------------------------------------------------------------
def fmt(x):
    if x >= 1e6: return f"{x/1e6:.2f}M"
    if x >= 1e3: return f"{x/1e3:.1f}k"
    return f"{x:.1f}"

print(f"{'#':<3}{'config':<30}{'mod':<6}{'MW':<4}{'FTT':>6}{'revs':>6}{'cells_M':>9}{'core_h':>10}{'wall_d':>8}{'E_MWh':>9}{'EUR':>11}")
for r in rows:
    print(f"{r['case']:<3}{r['config'][:29]:<30}{r['model']:<6}{r['size_MW']:<4}"
          f"{r['flow_throughs']:>6.2f}{r['rotor_revs']:>6.0f}{r['cells_M']:>9.1f}"
          f"{fmt(r['core_hours']):>10}{r['wall_days']:>8.2f}"
          f"{r['energy_MWh']:>9.2f}{r['cost_EUR']:>11,.0f}")
cs=[r['core_hours'] for r in rows]; eu=[r['cost_EUR'] for r in rows]
print(f"\ncore-hours: {min(cs):,.0f} .. {max(cs):,.0f}")
print(f"cost EUR:   {min(eu):,.0f} .. {max(eu):,.0f}")

with open("cases.json", "w") as f:
    json.dump(rows, f, indent=2)

# ----------------------------------------------------------------------
# 4. EXCEL OUTPUT
# ----------------------------------------------------------------------
wb = Workbook()

# ---- Sheet 1: Assumptions ----
ws = wb.active
ws.title = "Assumptions"
hdr = Font(bold=True, color="FFFFFF", name="Arial")
hdrfill = PatternFill("solid", fgColor="1F4E78")
bold = Font(bold=True, name="Arial")
ws["A1"] = "Computational-cost model - key assumptions"
ws["A1"].font = Font(bold=True, size=14, name="Arial")
arows = [
    ("", ""),
    ("HARDWARE - Leonardo (CINECA)", ""),
    ("DCGP node cores (2x Xeon 8480+)", DCGP_CORES_PER_NODE),
    ("DCGP node power draw (kW)", DCGP_NODE_KW),
    ("Booster node power draw (kW, 4x A100)", BOOSTER_NODE_KW),
    ("Booster node CPU-core equivalent (FV-CFD)", GPU_NODE_CORE_EQUIV),
    ("", ""),
    ("ENERGY / COST", ""),
    ("PUE (cooling overhead incl.)", PUE),
    ("Electricity price (EUR/kWh)", EUR_PER_KWH),
    ("", ""),
    ("SOLVER COST (core-seconds / cell / step)", ""),
    ("URANS", C_SOLVER["URANS"]),
    ("LES", C_SOLVER["LES"]),
    ("", ""),
    ("PER-TURBINE CELLS - 5 MW baseline (millions)", ""),
    ("Full geometry, URANS", PER_TURBINE[("FG","URANS")]),
    ("Full geometry, LES", PER_TURBINE[("FG","LES")]),
    ("Actuator line/disk, URANS", PER_TURBINE[("AL","URANS")]),
    ("Actuator line/disk, LES", PER_TURBINE[("AL","LES")]),
    ("", ""),
    ("SIZE FACTOR on per-turbine cells", ""),
    ("Full geometry: 5 MW / 15 MW", "1.0 / 1.8"),
    ("Actuator: 5 MW / 15 MW", "1.0 / 1.4"),
    ("", ""),
    ("ROTOR DIAMETER (m)", ""),
    ("NREL 5 MW", ROTOR_D[5]),
    ("IEA 15 MW", ROTOR_D[15]),
    ("", ""),
    ("MIN CELL in max-refinement zone", ""),
    ("Full geometry (wall-normal, y+~1): LES / URANS", "1e-5 m / 2e-5 m"),
    ("Actuator (rotor/wake box): LES / URANS", "D/100 / D/64"),
    ("", ""),
    ("WIND-FARM LAYOUT (aligned array, wind along columns)", ""),
    ("Wind farm 20: cols x rows", "5 x 4"),
    ("Wind farm 40: cols x rows", "8 x 5"),
    ("Spacing Sx (streamwise) x Sy (lateral), in D", "7 x 5"),
    ("Margins: upstream / downstream / sides (D)", "~6 / ~10-15 / ~5"),
    ("", ""),
    ("DOMAIN EXTENT Lx x Ly x Lz (in rotor diameters D)", ""),
    ("Single turbine", "24 x 8 x 6"),
    ("Wind farm 20", "44 x 26 x 6"),
    ("Wind farm 40", "64 x 32 x 6"),
    ("", ""),
    ("TIME / CONVERGENCE (all cases resolve the WAKE)", ""),
    ("Hub-height inflow U (m/s)", U_HUB),
    ("Rotor speed 5 MW / 15 MW (rpm)", "12 / 7.5"),
    ("Converge wake statistics over flow-through times", ""),
    ("   N flow-throughs LES (single/20/40)", "10 / 12 / 14"),
    ("   N flow-throughs URANS (single/20/40)", "6 / 8 / 10"),
    ("Time step - blade-resolved URANS", "2 deg azimuth"),
    ("Time step - blade-resolved LES", "1e-4 s (CFL)"),
    ("Time step - actuator (LES & URANS)", "tip transit: dx_min / V_tip"),
    ("", ""),
    ("Scenario", "Single sim, one wind speed, neutral ABL"),
]
r = 2
for k, v in arows:
    ws.cell(r, 1, k); ws.cell(r, 2, v)
    if k and v == "":
        ws.cell(r, 1).font = bold
    r += 1
ws.column_dimensions["A"].width = 46
ws.column_dimensions["B"].width = 38

# ---- Sheet 2: 24-case grid ----
ws2 = wb.create_sheet("24 cases")
cols = ["Case", "Configuration", "Turb. model", "Turbine (MW)", "Rotor D (m)",
        "# Turbines", "Layout (cols x rows)", "Spacing Sx x Sy (D)",
        "Domain LxLyLz (in D)", "Domain LxLyLz (km)",
        "Flow-through times", "Rotor revolutions", "Sim. time (s)",
        "Mesh (M cells)", "Min cell, max-refine zone (m)", "Min cell / D",
        "Time steps", "Core-hours", "Leonardo partition",
        "Node-hours", "# Nodes (job)", "Wall-clock (days)", "Energy (MWh)", "Cost (EUR)"]
ws2.append(cols)
for c in range(1, len(cols)+1):
    cell = ws2.cell(1, c); cell.font = hdr; cell.fill = hdrfill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
for r in rows:
    ws2.append([
        int(r["case"]), r["config"], r["model"], r["size_MW"], r["rotor_D_m"],
        r["n_turbines"], r["layout"], r["spacing_D"],
        r["domain_D"], r["domain_km"],
        round(r["flow_throughs"], 2), round(r["rotor_revs"]), round(r["T_sim_s"]),
        r["cells_M"], r["dx_min_m"], r["dx_over_D"],
        r["n_steps"], round(r["core_hours"]), r["partition"],
        round(r["node_hours"]), r["nodes"], round(r["wall_days"], 2),
        round(r["energy_MWh"], 2), round(r["cost_EUR"]),
    ])
widths = [6, 30, 11, 12, 11, 10, 16, 16, 16, 16, 13, 13, 12, 13, 16, 12, 11, 13, 17, 12, 12, 14, 12, 13]
for i, w in enumerate(widths, 1):
    ws2.column_dimensions[ws2.cell(1, i).column_letter].width = w
# number formats
for rr in range(2, len(rows)+2):
    ws2.cell(rr, 5).number_format = "#,##0"        # rotor D
    ws2.cell(rr, 11).number_format = "#,##0.00"    # flow-throughs
    ws2.cell(rr, 12).number_format = "#,##0"       # revolutions
    ws2.cell(rr, 13).number_format = "#,##0"       # sim time
    ws2.cell(rr, 14).number_format = "#,##0.0"     # mesh M
    ws2.cell(rr, 15).number_format = "0.00E+00"    # min cell (m)
    ws2.cell(rr, 16).number_format = "0.00E+00"    # min cell / D
    ws2.cell(rr, 17).number_format = "#,##0"       # steps
    ws2.cell(rr, 18).number_format = "#,##0"       # core-h
    ws2.cell(rr, 20).number_format = "#,##0"       # node-h
    ws2.cell(rr, 22).number_format = "#,##0.00"    # wall days
    ws2.cell(rr, 23).number_format = "#,##0.00"    # energy
    ws2.cell(rr, 24).number_format = "#,##0"       # cost
# zebra + color by model
thin = Side(style="thin", color="D9D9D9")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
for rr in range(2, len(rows)+2):
    model = ws2.cell(rr, 3).value
    fillcol = "EAF1F8" if model == "LES" else "FBEEE6"
    for cc in range(1, len(cols)+1):
        cell = ws2.cell(rr, cc)
        cell.fill = PatternFill("solid", fgColor=fillcol)
        cell.border = border
        cell.font = Font(name="Arial", size=10)
ws2.freeze_panes = "A2"
ws2.auto_filter.ref = f"A1:{ws2.cell(1, len(cols)).column_letter}{len(rows)+1}"

wb.save("wind_cfd_cost_grid.xlsx")
print("\nSaved wind_cfd_cost_grid.xlsx and cases.json")
